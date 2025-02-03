import os
import csv
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import undetected_chromedriver as uc


class CommentSize:
    def __init__(self):
        """
            Initializes the web driver session with optimized settings.
        """
        self.driver = self.setup_webdriver()

    @staticmethod
    def setup_webdriver():
        """
            Configures the undetected Chrome WebDriver with various options to:
            - Optimize performance
            - Avoid detection mechanisms
            - Improve page loading speed
        """

        options = uc.ChromeOptions()

        # Disable unnecessary features for better performance
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--process-per-site")
        options.add_argument("--log-level=3")
        options.add_argument("--remote-debugging-port=0")
        options.add_argument("--single-process")
        options.add_argument("--disable-background-timer-throttling")
        options.add_argument("--disable-backgrounding-occluded-windows")
        options.add_argument("--disable-renderer-backgrounding")
        options.add_argument("--disable-webgl")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-flash-plugin")
        options.add_argument("--dns-prefetch-disable")
        options.add_argument("--enable-quic")
        options.add_argument("--disable-webrtc")
        options.add_argument("--window-size=1280x800")

        # Disable loading images to improve speed
        options.add_experimental_option("prefs", {
            "profile.managed_default_content_settings.images": 2,
        })

        # Set WebDriver capabilities to optimize page loading
        options.page_load_strategy = 'eager'

        # Initializing browser session
        webdriver = uc.Chrome(options=options)

        # Blocking unnecessary network requests (images, fonts, videos, etc.) to speed up loading
        webdriver.execute_cdp_cmd("Network.enable", {})
        webdriver.execute_cdp_cmd("Network.setBlockedURLs", {
            "urls": ["*.png", "*.jpg", "*.jpeg", "*.gif", "*.svg", "*.woff", "*.woff2", "*.ttf", "*.ico",  # Fonts
                     "*.mp4", "*.webm", "*.avi", "*.mov", "*.mkv",
                     "*.json", "*.xml"]})

        return webdriver

    def fetch_comment_count(self, url):
        """
        Extract the number of comments from a given URL using Selenium.
        """
        try:
            self.driver.get(url)
            element = self.driver.find_element(By.TAG_NAME, "data")
            return element.get_attribute("data-value") if element else "N/A"
        except Exception as e:
            print(f"Error processing {url}: {e}")
            return "Error"

    def process_urls(self, filepath, start_row, end_row, url_column, index_to_place_nums):
        """
            Extract and update comment counts for URLs stored in a DataFrame.
            Adds a new column 'no. of comments' instead of updating an existing column.
        """
        if not os.path.exists(filepath):
            print(f"File {filepath} not found!")
            return

        try:
            if filepath.endswith(".xlsx"):
                wb = load_workbook(filepath)
                sheet = wb.active

                url_col_index = url_column
                comment_num_index = index_to_place_nums

                if sheet.cell(row=1, column=comment_num_index).value != "No. of Comments":
                    sheet.cell(row=1, column=comment_num_index).value = "No. of Comments"

                for row in range(start_row, min(end_row + 1, sheet.max_row + 1)):
                    url = sheet.cell(row=row, column=url_col_index).value
                    comment_count = sheet.cell(row=row, column=comment_num_index).value

                    if not url or (comment_count and comment_count != ""):
                        print(f"Skipping row {row}, already processed or no URL found.")
                        continue

                    print(f"Processing row {row}: {url}")
                    sheet.cell(row=row, column=comment_num_index, value=self.fetch_comment_count(url))

                    try:
                        wb.save(filepath)
                        print(f"File successfully saved: {filepath}")  # Debugging output
                    except Exception as e:
                        print(f"Error saving file: {e}")

                wb.close()
                print("Processing complete. Data saved.")

            elif filepath.endswith(".csv"):
                with open(filepath, mode="r") as file:
                    reader = csv.reader(file)
                    header = next(reader)
                    rows = list(reader)

                # Add "No. of Comments" column if missing
                if "No. of Comments" not in header:
                    header.append("No. of Comments")

                comment_col_index = header.index("No. of Comments")

                for index in range(start_row - 2, min(end_row - 1, len(rows))):
                    row = rows[index]
                    url = row[url_column - 1] if url_column - 1 < len(row) else None
                    existing_comment = row[comment_col_index] if comment_col_index < len(row) else ""

                    if not url or (existing_comment and existing_comment != ""):
                        print(f"Skipping row {index + 2}, already processed or no URL found.")
                        continue

                    print(f"Processing row {index + 2}: {url}")
                    comment_count = self.fetch_comment_count(url)

                    # Ensure the row is long enough
                    if len(row) <= comment_col_index:
                        row.extend([""] * (comment_col_index - len(row) + 1))

                    row[comment_col_index] = comment_count

                    with open(filepath, mode="w", newline="") as file:
                        writer = csv.writer(file)
                        writer.writerow(header)
                        writer.writerows(rows)

                print("Processing complete. Data saved.")

            else:
                raise ValueError("Unsupported file format. Use .xlsx or .csv")

        except KeyboardInterrupt:
            print("Process interrupted. Saving progress...")


if __name__ == "__main__":
    input_file = "scraped_data.csv"
    scraper = CommentSize()
    scraper.process_urls(input_file, start_row=2, end_row=10, url_column=3, index_to_place_nums=6)


