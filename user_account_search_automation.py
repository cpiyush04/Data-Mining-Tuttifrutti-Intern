import csv
import os
import time
from openpyxl import load_workbook
from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import undetected_chromedriver as uc


class SocialMediaProfileScraper:
    def __init__(self):
        """
            Initializes the web driver session with optimized settings.
        """
        self.driver = self.setup_webdriver()

    @staticmethod
    def setup_webdriver():
        """
            Configures the undetected Chrome WebDriver with various options to:
            - Optimize performance
            - Avoid detection mechanisms
            - Improve page loading speed
        """

        options = uc.ChromeOptions()

        # Disable unnecessary features for better performance
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--process-per-site")
        options.add_argument("--log-level=3")
        options.add_argument("--remote-debugging-port=0")
        options.add_argument("--single-process")
        options.add_argument("--disable-background-timer-throttling")
        options.add_argument("--disable-backgrounding-occluded-windows")
        options.add_argument("--disable-renderer-backgrounding")
        options.add_argument("--disable-webgl")
        options.add_argument("--disable-plugins")
        options.add_argument("--disable-flash-plugin")
        options.add_argument("--dns-prefetch-disable")
        options.add_argument("--enable-quic")
        options.add_argument("--disable-webrtc")
        options.add_argument("--window-size=1280x800")

        # Disable loading images and cookies to improve speed
        options.add_experimental_option("prefs", {
            "profile.managed_default_content_settings.cookies": 2,
            "profile.managed_default_content_settings.images": 2,
        })

        # Set WebDriver capabilities to optimize page loading
        options.page_load_strategy = 'eager'

        # Initializing browser session
        webdriver = uc.Chrome(options=options)

        # Blocking unnecessary network requests (images, fonts, videos, etc.) to speed up loading
        webdriver.execute_cdp_cmd("Network.enable", {})
        webdriver.execute_cdp_cmd("Network.setBlockedURLs", {
            "urls": ["*.png", "*.jpg", "*.jpeg", "*.gif", "*.svg", "*.woff", "*.woff2", "*.ttf", "*.ico",  # Fonts
                     "*.mp4", "*.webm", "*.avi", "*.mov", "*.mkv",
                     "*.json", "*.xml"]})

        # Open Google as the default page
        webdriver.get("https://www.google.com")
        return webdriver

    @staticmethod
    def load_commentator_names(filepath, start_row, end_row, name_col_idx):
        """
            Reads a list of usernames from an Excel (.xlsx) or CSV (.csv) file.

            Parameters:
            - filepath: Path to the input file
            - start_row: Row number to start reading from (1-based index)
            - end_row: Row number to stop reading at (inclusive)
            - name_col_idx: Column index where usernames are stored (1-based index)

            Returns:
            - A list of usernames extracted from the file.
        """
        usernames = []

        # Process Excel files
        if filepath.endswith(".xlsx"):
            wb = load_workbook(filename=filepath)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=start_row, max_row=end_row, max_col=name_col_idx, values_only=True):
                if row[name_col_idx - 1]:
                    usernames.append(row[1])
            return usernames

        # Process CSV files
        elif filepath.endswith(".csv"):
            with open(filepath, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                for idx, row in enumerate(reader, start=1):
                    if start_row <= idx <= end_row and len(row) > 1 and row[1]:
                        usernames.append(row[1])
                return usernames

    def profile_search(self, username):
        """
            Searches for social media profiles related to the given username using Google.

            Parameters:
            - username: The username to search for

            Returns:
            - A list of URLs that potentially link to the user's social media profiles.
        """

        try:
            self.driver.delete_all_cookies()
            search_query = f'"{username}" ("gamer" OR "streamer") Instagram OR ' \
                           f'LinkedIn OR Twitter OR Facebook OR YouTube'
            wait = WebDriverWait(self.driver, 10)
            search_box = wait.until(EC.presence_of_element_located((By.NAME, "q")))
            search_box.clear()
            search_box.send_keys(search_query)
            search_box.submit()
            profile_urls = []
            time.sleep(3)
            results = self.driver.find_elements(By.CSS_SELECTOR, 'a[jsname="UWckNb"]')
            for result in results:
                link = result.get_attribute("href")
                if link:
                    profile_urls.append(link)
            return profile_urls

        except Exception:
            print(f"Error while searching for {username}")
            return []  # Return an empty list if search fails

    def process_save_output(self, usernames, filepath):
        """
            Processes the list of usernames and saves their social media profile links to a CSV file.

            Parameters:
            - usernames: List of usernames to process
            - filepath: Path to the output CSV file

            Saves:
            - A CSV file containing usernames and their corresponding social media profile links.
        """

        file_exists = os.path.exists(filepath)

        # Open the output file in append mode
        with open(filepath, "a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if not file_exists:
                writer.writerow(["Profile Name"] + [f"{i + 1}" for i in range(9)])

            try:
                for username in usernames:
                    links = self.profile_search(username)
                    if links:
                        writer.writerow([username] + links[:9])  # Writing without reopening the file
                    else:
                        print(f"Could not fetch {username}\n")
                        break
            except KeyboardInterrupt:
                print("Process interrupted. Saved Progress")
            finally:
                self.driver.quit()
                print(f"Driver Closed")


if __name__ == "__main__":
    input_file = "games_and_commentators.xlsx"
    output_file = "output.csv"
    scraper = SocialMediaProfileScraper()
    users_list = scraper.load_commentator_names(input_file, start_row=2, end_row=6, name_col_idx=2)
    scraper.process_save_output(users_list, output_file)
